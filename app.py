"""
CPK Analyzer - 미니탭 스타일 공정능력 분석기
Usage: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
from scipy import stats
from scipy.special import gamma as _gamma
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
from io import BytesIO
import zipfile
import xml.etree.ElementTree as ET
import base64

# --------------------------------------------------
# Monkey-patch openpyxl (일부 엑셀 파일 호환성 문제 해결)
# --------------------------------------------------
try:
    from openpyxl.packaging.custom import CustomPropertyList
    _orig_from_tree = CustomPropertyList.from_tree.__func__

    @classmethod
    def _safe_from_tree(cls, tree):
        try:
            return _orig_from_tree(cls, tree)
        except TypeError:
            return cls()

    CustomPropertyList.from_tree = _safe_from_tree
except Exception:
    pass

# --------------------------------------------------
# 페이지 설정
# --------------------------------------------------
st.set_page_config(page_title="CPK 분석기", layout="wide")

st.markdown("""
<style>
    .block-container { padding-top: 1rem; }
</style>
""", unsafe_allow_html=True)


# ==========================================================
#  1. 데이터 파싱
# ==========================================================

def get_sheet_names(file_bytes: bytes) -> list[str]:
    """엑셀 파일에서 시트 이름 목록을 가져온다."""
    try:
        xls = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
        return xls.sheet_names
    except Exception:
        z = zipfile.ZipFile(BytesIO(file_bytes))
        tree = ET.parse(z.open("xl/workbook.xml"))
        ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        return [s.attrib["name"] for s in tree.findall(".//m:sheet", ns)]


def parse_sheet(file_bytes: bytes, sheet_name: str):
    """시트를 파싱하여 테스트 항목 리스트를 반환한다."""
    df = pd.read_excel(
        BytesIO(file_bytes), sheet_name=sheet_name,
        header=None, engine="openpyxl",
    )

    header_row = header_col = None
    for i in range(min(15, len(df))):
        for j in range(min(5, len(df.columns))):
            val = str(df.iloc[i, j]).strip() if pd.notna(df.iloc[i, j]) else ""
            if "측정항목" in val:
                header_row, header_col = i, j
                break
        if header_row is not None:
            break
    if header_row is None:
        return None, "'측정항목' 행을 찾을 수 없습니다."

    spec_max_row = spec_min_row = None
    for i in range(header_row + 1, min(header_row + 12, len(df))):
        cell = str(df.iloc[i, header_col]).strip().lower() if pd.notna(df.iloc[i, header_col]) else ""
        if spec_max_row is None and ("spec max" in cell or cell == "max"):
            spec_max_row = i
        elif spec_min_row is None and ("spec min" in cell or cell == "min"):
            spec_min_row = i
    if spec_max_row is None or spec_min_row is None:
        return None, "Spec max / Spec min 행을 찾을 수 없습니다."

    result_col = None
    for j in range(header_col + 1, len(df.columns)):
        val = str(df.iloc[header_row, j]).strip().lower() if pd.notna(df.iloc[header_row, j]) else ""
        if "result" in val:
            result_col = j
            break
    if result_col is None:
        result_col = len(df.columns)

    data_start_row = None
    for i in range(header_row + 1, len(df)):
        val = str(df.iloc[i, header_col]).strip() if pd.notna(df.iloc[i, header_col]) else ""
        if len(val) > 10 and any(c.isdigit() for c in val) and any(c.isalpha() for c in val):
            data_start_row = i
            break
    if data_start_row is None:
        return None, "시리얼번호 행을 찾을 수 없습니다."

    tests = []
    skipped = []
    for col in range(header_col + 1, result_col):
        name = str(df.iloc[header_row, col]).strip() if pd.notna(df.iloc[header_row, col]) else ""
        if not name or name == "nan":
            continue
        raw_max = df.iloc[spec_max_row, col]
        raw_min = df.iloc[spec_min_row, col]
        try:
            usl = float(raw_max)
            lsl = float(raw_min)
        except (ValueError, TypeError):
            skipped.append(f"{name} (Spec 변환 실패: max={raw_max}, min={raw_min})")
            continue
        values = []
        for i in range(data_start_row, len(df)):
            v = df.iloc[i, col]
            if pd.notna(v):
                try:
                    values.append(float(v))
                except (ValueError, TypeError):
                    pass
        if len(values) < 2:
            skipped.append(f"{name} (데이터 부족: {len(values)}개)")
            continue
        tests.append({"name": name, "usl": usl, "lsl": lsl, "data": np.array(values)})

    if not tests:
        msg = "분석 가능한 테스트 항목이 없습니다."
        if skipped:
            msg += "\n건너뛴 항목: " + ", ".join(skipped)
        return None, msg
    return tests, skipped if skipped else None


# ==========================================================
#  2. 공정능력 계산 (미니탭 동일 방식)
# ==========================================================

def calculate_capability(data: np.ndarray, usl: float, lsl: float) -> dict:
    n = len(data)
    mean = np.mean(data)
    target = (usl + lsl) / 2.0

    stdev_overall = np.std(data, ddof=1)

    # Within StDev (s/c4 방식 — 미니탭 결과와 정확히 일치)
    c4 = np.sqrt(2.0 / (n - 1)) * _gamma(n / 2.0) / _gamma((n - 1) / 2.0)
    stdev_within = stdev_overall / c4

    cp  = (usl - lsl) / (6 * stdev_within)
    cpl = (mean - lsl) / (3 * stdev_within)
    cpu = (usl - mean) / (3 * stdev_within)
    cpk = min(cpl, cpu)

    pp  = (usl - lsl) / (6 * stdev_overall)
    ppl = (mean - lsl) / (3 * stdev_overall)
    ppu = (usl - mean) / (3 * stdev_overall)
    ppk = min(ppl, ppu)

    tau = np.sqrt(stdev_overall ** 2 + (mean - target) ** 2)
    cpm = (usl - lsl) / (6 * tau)

    ppm_lsl_obs = (np.sum(data < lsl) / n) * 1e6
    ppm_usl_obs = (np.sum(data > usl) / n) * 1e6
    ppm_total_obs = ppm_lsl_obs + ppm_usl_obs

    ppm_lsl_w = stats.norm.cdf((lsl - mean) / stdev_within) * 1e6
    ppm_usl_w = (1 - stats.norm.cdf((usl - mean) / stdev_within)) * 1e6
    ppm_total_w = ppm_lsl_w + ppm_usl_w

    ppm_lsl_o = stats.norm.cdf((lsl - mean) / stdev_overall) * 1e6
    ppm_usl_o = (1 - stats.norm.cdf((usl - mean) / stdev_overall)) * 1e6
    ppm_total_o = ppm_lsl_o + ppm_usl_o

    return dict(
        n=n, mean=mean, target=target, lsl=lsl, usl=usl,
        stdev_within=stdev_within, stdev_overall=stdev_overall,
        cp=cp, cpl=cpl, cpu=cpu, cpk=cpk,
        pp=pp, ppl=ppl, ppu=ppu, ppk=ppk, cpm=cpm,
        ppm_lsl_obs=ppm_lsl_obs, ppm_usl_obs=ppm_usl_obs, ppm_total_obs=ppm_total_obs,
        ppm_lsl_w=ppm_lsl_w, ppm_usl_w=ppm_usl_w, ppm_total_w=ppm_total_w,
        ppm_lsl_o=ppm_lsl_o, ppm_usl_o=ppm_usl_o, ppm_total_o=ppm_total_o,
    )


# ==========================================================
#  3. 미니탭 스타일 차트 (Option B 레이아웃)
# ==========================================================

def _draw_table(fig, x, y, title, rows, width=0.28, fs=9, rh=0.023):
    """figure 좌표계에 테두리 있는 표를 그린다."""
    n_rows = len(rows)
    title_h = rh * 1.3
    pad = rh * 0.35
    total_h = title_h + rh * n_rows + pad

    rect = Rectangle(
        (x, y - total_h), width, total_h,
        fill=True, facecolor="white", edgecolor="black",
        linewidth=1.0, transform=fig.transFigure, zorder=5,
    )
    fig.patches.append(rect)

    fig.text(
        x + width / 2, y - title_h * 0.45, title,
        fontsize=fs, fontweight="bold", ha="center", va="center",
        transform=fig.transFigure, zorder=6, family="sans-serif",
    )

    line = Line2D(
        [x, x + width], [y - title_h, y - title_h],
        color="black", linewidth=0.5,
        transform=fig.transFigure, zorder=6,
    )
    fig.lines.append(line)

    for i, (label, value) in enumerate(rows):
        ry = y - title_h - rh * (i + 0.6)
        fig.text(
            x + 0.010, ry, label,
            fontsize=fs - 0.5, ha="left", va="center",
            transform=fig.transFigure, zorder=6, family="sans-serif",
        )
        fig.text(
            x + width - 0.010, ry, str(value),
            fontsize=fs - 0.5, ha="right", va="center",
            transform=fig.transFigure, zorder=6, family="sans-serif",
        )

    return total_h


def create_capability_chart(test_name: str, data: np.ndarray,
                            usl: float, lsl: float, r: dict) -> plt.Figure:
    """Option B: Process Data 좌측 | 히스토그램 중앙 | Capability 우측"""

    fig = plt.figure(figsize=(14, 7.5))
    fig.patch.set_facecolor("#F0ECD8")

    mean = r["mean"]
    target = r["target"]
    sw = r["stdev_within"]
    so = r["stdev_overall"]

    # ── 제목 ──
    fig.text(
        0.50, 0.965, test_name,
        fontsize=13, fontweight="bold", ha="center", va="top",
        family="sans-serif",
    )

    # ── Process Data — 좌측 패널 ──
    _draw_table(
        fig, 0.02, 0.87,
        "Process Data",
        [
            ("LSL",            f"{lsl}"),
            ("Target",         f"{target:.5g}"),
            ("USL",            f"{usl}"),
            ("Sample Mean",    f"{mean:.5g}"),
            ("Sample N",       f"{r['n']}"),
            ("StDev(Within)",  f"{sw:.7g}"),
            ("StDev(Overall)", f"{so:.7g}"),
        ],
        width=0.17, fs=9.5, rh=0.028,
    )

    # ── 히스토그램 — 중앙 ──
    ax = fig.add_axes([0.215, 0.20, 0.42, 0.67])
    ax.set_facecolor("white")

    spec_range = usl - lsl
    sigma_max = max(sw, so)
    x_lo = min(lsl - spec_range * 0.08, mean - 4.5 * sigma_max)
    x_hi = max(usl + spec_range * 0.08, mean + 4.5 * sigma_max)

    n_bins = max(int(np.sqrt(len(data))), 10)
    ax.hist(
        data, bins=n_bins, density=True,
        color="#787878", edgecolor="black", linewidth=0.8,
        alpha=1.0, zorder=3,
    )

    x_curve = np.linspace(x_lo, x_hi, 500)
    ax.plot(x_curve, stats.norm.pdf(x_curve, mean, sw),
            "r-", linewidth=1.8, zorder=4)
    ax.plot(x_curve, stats.norm.pdf(x_curve, mean, so),
            "k--", linewidth=1.8, zorder=4)

    ax.axvline(lsl,    color="red",   linestyle="--", linewidth=1.5, zorder=2)
    ax.axvline(usl,    color="red",   linestyle="--", linewidth=1.5, zorder=2)
    ax.axvline(target, color="green", linestyle="--", linewidth=1.5, zorder=2)

    yt = ax.get_ylim()[1]
    ax.text(lsl,    yt * 1.04, "LSL",    color="red",   ha="center",
            fontsize=9, fontweight="bold", clip_on=False)
    ax.text(target, yt * 1.04, "Target", color="green", ha="center",
            fontsize=9, fontweight="bold", clip_on=False)
    ax.text(usl,    yt * 1.04, "USL",    color="red",   ha="center",
            fontsize=9, fontweight="bold", clip_on=False)

    ax.set_xlim(x_lo, x_hi)
    ax.set_ylim(bottom=0)
    ax.yaxis.set_visible(False)
    ax.tick_params(axis="x", labelsize=9)

    # ── 범례 (히스토그램 내부) ──
    legend_handles = [
        Line2D([0], [0], color="red",   linewidth=2, linestyle="-",  label="Within"),
        Line2D([0], [0], color="black", linewidth=2, linestyle="--", label="Overall"),
    ]
    ax.legend(handles=legend_handles, loc="upper right", fontsize=9,
              frameon=True, fancybox=False, edgecolor="black")

    # ── Capability 테이블 — 우측 ──
    _draw_table(
        fig, 0.68, 0.84,
        "Potential (Within) Capability",
        [("Cp", f"{r['cp']:.2f}"), ("CPL", f"{r['cpl']:.2f}"),
         ("CPU", f"{r['cpu']:.2f}"), ("Cpk", f"{r['cpk']:.2f}")],
        width=0.29,
    )

    _draw_table(
        fig, 0.68, 0.645,
        "Overall Capability",
        [("Pp", f"{r['pp']:.2f}"), ("PPL", f"{r['ppl']:.2f}"),
         ("PPU", f"{r['ppu']:.2f}"), ("Ppk", f"{r['ppk']:.2f}"),
         ("Cpm", f"{r['cpm']:.2f}")],
        width=0.29,
    )

    # ── 하단 PPM 테이블 ──
    btm_w = 0.29
    _draw_table(
        fig, 0.04, 0.15, "Observed Performance",
        [("PPM < LSL", f"{r['ppm_lsl_obs']:.2f}"),
         ("PPM > USL", f"{r['ppm_usl_obs']:.2f}"),
         ("PPM Total", f"{r['ppm_total_obs']:.2f}")],
        width=btm_w, fs=8,
    )
    _draw_table(
        fig, 0.35, 0.15, "Exp. Within Performance",
        [("PPM < LSL", f"{r['ppm_lsl_w']:.2f}"),
         ("PPM > USL", f"{r['ppm_usl_w']:.2f}"),
         ("PPM Total", f"{r['ppm_total_w']:.2f}")],
        width=btm_w, fs=8,
    )
    _draw_table(
        fig, 0.66, 0.15, "Exp. Overall Performance",
        [("PPM < LSL", f"{r['ppm_lsl_o']:.2f}"),
         ("PPM > USL", f"{r['ppm_usl_o']:.2f}"),
         ("PPM Total", f"{r['ppm_total_o']:.2f}")],
        width=btm_w, fs=8,
    )

    plt.close(fig)
    return fig


def fig_to_png(fig, dpi=150) -> bytes:
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0)
    return buf.read()


# ==========================================================
#  3-1. 엑셀에 차트 삽입
# ==========================================================

def insert_charts_to_excel(file_bytes: bytes, chart_pngs: list,
                           test_names: list) -> tuple:
    """생성된 차트를 엑셀 CPK 시트의 '공정능력 분포도' 영역에 자동 삽입한다.

    원본 양식(병합 셀)을 유지하면서 열 너비를 균등화하고,
    각 슬롯에 맞는 크기로 차트를 삽입한다.
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.utils import get_column_letter

    wb = load_workbook(BytesIO(file_bytes))

    # ── 1) CPK 시트 찾기 ──
    ws = None
    for name in wb.sheetnames:
        if "cpk" in name.lower() or "출력" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        return None, "CPK 시트를 찾을 수 없습니다."

    # ── 2) "공정능력 분포도" 헤더 찾기 (가장 오른쪽 매치) ──
    matches = []
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and "공정능력" in str(cell.value):
                matches.append((cell.row, cell.column))
    if not matches:
        return None, "'공정능력 분포도' 텍스트를 찾을 수 없습니다."
    hdr_row, chart_start_col = max(matches, key=lambda x: x[1])

    # ── 3) 차트 영역 열 수 (병합 헤더 범위) ──
    n_cols = 1
    for mc in ws.merged_cells.ranges:
        if mc.min_row == hdr_row and mc.min_col == chart_start_col:
            n_cols = mc.max_col - mc.min_col + 1
            break

    # ── 차트 행 수 자동 감지 (병합 셀에서 추출) ──
    CHART_ROWS = 12  # 기본값
    first_chart_row = hdr_row + 2  # 헤더 → 타이틀 → 첫 차트 행
    for mc in ws.merged_cells.ranges:
        if (mc.min_col == chart_start_col
                and mc.min_row == first_chart_row):
            CHART_ROWS = mc.max_row - mc.min_row + 1
            break

    # ── 4) 열 너비 균등화 (가장 넓은 열 기준) ──
    widths = []
    for c in range(n_cols):
        letter = get_column_letter(chart_start_col + c)
        w = ws.column_dimensions[letter].width
        widths.append(w if w is not None else 8.43)

    target_w = max(widths)
    if target_w < 25:
        target_w = 25
    for c in range(n_cols):
        letter = get_column_letter(chart_start_col + c)
        ws.column_dimensions[letter].width = target_w

    # ── 5) 기존 이미지 제거 (차트 영역) ──
    n_groups = -(-len(chart_pngs) // n_cols)  # ceil
    max_row_clear = hdr_row + n_groups * (1 + CHART_ROWS) + 5
    col_lo = chart_start_col - 1   # 0-based
    col_hi = col_lo + n_cols
    keep = []
    for img in ws._images:
        try:
            a = img.anchor
            if hasattr(a, "_from"):
                if (col_lo <= a._from.col < col_hi
                        and hdr_row - 1 <= a._from.row < max_row_clear):
                    continue
        except Exception:
            pass
        keep.append(img)
    ws._images = keep

    # ── 6) 차트 삽입 — TwoCellAnchor로 셀에 정확히 맞춤 ──
    for i, (png, name) in enumerate(zip(chart_pngs, test_names)):
        col_off = i % n_cols
        group = i // n_cols
        col = chart_start_col + col_off
        title_row = hdr_row + 1 + group * (1 + CHART_ROWS)
        chart_row = title_row + 1

        # 타이틀
        ws.cell(row=title_row, column=col, value=name)

        # 이미지 — TwoCellAnchor + 패딩으로 셀 안에 여유 있게 배치
        PAD = 57150   # 약 6px 여백 (1px = 9525 EMU)
        img = XlImage(BytesIO(png))
        anchor = TwoCellAnchor()
        anchor._from = AnchorMarker(
            col=col - 1, colOff=PAD,        # 왼쪽 여백
            row=chart_row - 1, rowOff=PAD,   # 위쪽 여백
        )
        anchor.to = AnchorMarker(
            col=col, colOff=-PAD,            # 오른쪽 여백
            row=chart_row - 1 + CHART_ROWS, rowOff=-PAD,  # 아래쪽 여백
        )
        img.anchor = anchor
        ws.add_image(img)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), None


# ==========================================================
#  4. 클릭 복사 (Streamlit 네이티브 — PyInstaller 호환)
# ==========================================================

# 페이지 로드 시 한 번만 삽입되는 클립보드 복사 JS
COPY_JS = """
<script>
document.addEventListener('click', async function(e) {
    const img = e.target.closest('img[data-copyable]');
    if (!img) return;
    try {
        const c = document.createElement('canvas');
        c.width = img.naturalWidth;
        c.height = img.naturalHeight;
        c.getContext('2d').drawImage(img, 0, 0);
        const blob = await new Promise(r => c.toBlob(r, 'image/png'));
        await navigator.clipboard.write([
            new ClipboardItem({ 'image/png': blob })
        ]);
        // 복사 완료 피드백
        const old = img.style.outline;
        img.style.outline = '3px solid #0a7';
        img.title = 'Copied!';
        setTimeout(() => { img.style.outline = old; img.title = 'Click to copy'; }, 1200);
    } catch(err) {
        alert('복사 실패 — Chrome 또는 Edge에서 시도해주세요.');
    }
});
</script>
"""


def render_copyable_image(png_bytes: bytes, test_name: str):
    """st.image + data-copyable 속성으로 클릭 복사 지원."""
    b64 = base64.b64encode(png_bytes).decode()
    html_img = (
        f'<img src="data:image/png;base64,{b64}" '
        f'data-copyable="1" title="Click to copy" '
        f'style="width:100%;cursor:pointer;border-radius:4px;'
        f'border:1px solid #e0e0e0;" />'
    )
    st.markdown(html_img, unsafe_allow_html=True)


# ==========================================================
#  5. Streamlit 앱
# ==========================================================

def main():
    st.title("CPK Analyzer")
    st.caption("엑셀 파일을 업로드하면 미니탭 스타일의 공정능력 그래프를 자동 생성합니다.")

    # ── 사이드바 ──
    with st.sidebar:
        st.header("설정")
        uploaded = st.file_uploader("엑셀 파일 업로드", type=["xlsx", "xls"])

        if uploaded is None:
            st.info("엑셀 파일을 업로드하세요.")
            st.session_state.pop("cpk", None)
            st.session_state.pop("excel_out", None)
            return

        file_bytes = uploaded.getvalue()
        sheet_names = get_sheet_names(file_bytes)
        selected_sheet = st.selectbox("시트 선택", sheet_names)
        cols_per_row = st.radio("그래프 배치", [1, 2], index=1, horizontal=True)
        run = st.button("분석 시작", type="primary", use_container_width=True)

    # ── 분석 실행 ──
    if run:
        with st.spinner("데이터 분석 중..."):
            tests, info = parse_sheet(file_bytes, selected_sheet)

        if tests is None:
            st.error(info)
            return

        chart_pngs = []
        summary_rows = []
        for t in tests:
            r = calculate_capability(t["data"], t["usl"], t["lsl"])
            fig = create_capability_chart(t["name"], t["data"], t["usl"], t["lsl"], r)
            chart_pngs.append(fig_to_png(fig))
            summary_rows.append({
                "테스트": t["name"][:55],
                "N": r["n"],
                "평균": f"{r['mean']:.4f}",
                "Cp": f"{r['cp']:.2f}",
                "Cpk": f"{r['cpk']:.2f}",
                "Pp": f"{r['pp']:.2f}",
                "Ppk": f"{r['ppk']:.2f}",
                "판정": "PASS" if r["cpk"] >= 1.33 else "FAIL",
            })

        st.session_state["cpk"] = {
            "names": [t["name"] for t in tests],
            "pngs": chart_pngs,
            "summary": summary_rows,
            "info": info,
            "cols": cols_per_row,
        }
        st.session_state.pop("excel_out", None)

    # ── 결과 표시 ──
    if "cpk" not in st.session_state:
        st.info("사이드바에서 파일을 업로드하고 '분석 시작'을 누르세요.")
        return

    ctx = st.session_state["cpk"]
    names = ctx["names"]
    chart_pngs = ctx["pngs"]
    summary_rows = ctx["summary"]
    info = ctx["info"]
    cpr = ctx["cols"]

    if info:
        with st.expander(f"건너뛴 항목 ({len(info)}개)", expanded=False):
            for s in info:
                st.warning(s)

    # ── 요약 ──
    n_pass = sum(1 for s in summary_rows if s["판정"] == "PASS")
    n_fail = len(summary_rows) - n_pass
    col_a, col_b, col_c = st.columns(3)
    col_a.metric("전체 항목", f"{len(names)}개")
    col_b.metric("PASS", f"{n_pass}개")
    col_c.metric("FAIL", f"{n_fail}개")

    summary_df = pd.DataFrame(summary_rows)
    st.dataframe(
        summary_df.style.map(
            lambda v: "color: #0a7; font-weight: bold" if v == "PASS"
            else ("color: #d33; font-weight: bold" if v == "FAIL" else ""),
            subset=["판정"],
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.divider()
    st.caption("차트를 클릭하면 클립보드에 복사됩니다. Ctrl+V로 붙여넣기 하세요.")

    # 클립보드 복사 JS 삽입 (한 번만)
    st.markdown(COPY_JS, unsafe_allow_html=True)

    # ── 그래프 표시 ──
    for row_start in range(0, len(names), cpr):
        grid = st.columns(cpr)
        for col_idx in range(cpr):
            i = row_start + col_idx
            if i >= len(names):
                break
            with grid[col_idx]:
                render_copyable_image(chart_pngs[i], names[i])

    # ── 엑셀 차트 삽입 ──
    st.divider()
    st.subheader("엑셀에 차트 삽입")
    st.caption("업로드한 엑셀의 CPK 시트 '공정능력 분포도' 영역에 차트를 자동 삽입합니다.")

    if st.button("차트 삽입 실행", type="primary"):
        with st.spinner("엑셀 파일 처리 중... (대용량 파일은 시간이 걸릴 수 있습니다)"):
            result, err = insert_charts_to_excel(file_bytes, chart_pngs, names)
        if err:
            st.error(err)
        else:
            st.session_state["excel_out"] = result
            st.success("차트 삽입 완료!")

    if "excel_out" in st.session_state:
        out_name = uploaded.name if uploaded else "CPK_result.xlsx"
        st.download_button(
            "수정된 엑셀 다운로드",
            st.session_state["excel_out"],
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
